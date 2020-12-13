from flask import Flask, request, jsonify
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

BUFFER_SIZE = 1
buf = []

def makeTable(sheet):
	sheet.column_dimensions['A'].width = 8
	sheet.column_dimensions['B'].width = 70
	sheet.column_dimensions['C'].width = 27
	sheet.column_dimensions['D'].width = 32
	sheet.column_dimensions['E'].width = 12
	sheet['A1'] = "N"
	sheet['B1'] = "User ID"
	sheet['C1'] = "Datetime"
	sheet['D1'] = "Item"
	sheet['E1'] = "Prise"
	for i in range(5):
		sheet[1][i].fill = PatternFill(start_color='F0F8F8F8',
                   end_color='F0F8F8F8',
                   fill_type='solid')
		sheet[1][i].alignment = Alignment(horizontal='center', vertical='center')
		sheet[1][i].font = Font(bold=True)
	return sheet

def getIdEmptyCell(sheet):
	i = 1
	while not (sheet[i][3].value is None):
		i += 1
	return i

#borders and Aligment
def setMainStyle(sheet, y, x):
	bstyle = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
	for iy in range (1, y):
		for ix in range (0, x):
			sheet[iy][ix].border = bstyle
			sheet[iy][ix].alignment = Alignment(horizontal='center', vertical='center')
	return sheet

def getPrevID(sheet, y):
	i = y-1
	while (i > 1) and (sheet[i][0].value is None):
		i -= 1
	if i <= 1:
		return 1
	else:
		return sheet[i][0].value+1

def saveJson(sheet):
	y = getIdEmptyCell(sheet)
	for i in range(len(buf)):
		sheet[y][0].value = 1 if y <= 1 else getPrevID(sheet, y) 
		sheet[y][1].value = buf[i]["user_id"]
		sheet[y][2].value = buf[i]["datetime"]
		oldy = y
		for item in buf[i]["check"]:
			sheet[y][3].value = item["item"]
			sheet[y][4].value = item["price"]
			y += 1
		sheet.merge_cells(start_row=oldy, start_column=1, end_row=y-1, end_column=1)
		sheet.merge_cells(start_row=oldy, start_column=2, end_row=y-1, end_column=2)
		sheet.merge_cells(start_row=oldy, start_column=3, end_row=y-1, end_column=3)
	sheet = setMainStyle(sheet, y, 5)
	return sheet

def saveBuffer():
	global buf
	try:
		book = openpyxl.open("data.xlsx", read_only=False)
	except:
		book = openpyxl.Workbook()
	sheet = book.active
	if sheet['A1'].value is None:
		sheet = makeTable(sheet)
	sheet = saveJson(sheet)
	book.save("data.xlsx")


def writeData(content):
	global buf
	buf.append(content);
	buf[-1]["datetime"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
	if len(buf) >= BUFFER_SIZE:
		saveBuffer()
		buf = []


app = Flask(__name__)
 
@app.route('/', methods=['POST', 'GET'])
def index():
	if request.is_json:
		content = request.get_json()
		writeData(content)
		return 'OK'
 
if __name__ == "__main__":
	app.run()